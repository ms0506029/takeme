# -*- coding: utf-8 -*-
"""
pro_cli.py — Purchase Control Tower PRO CLI
功能：
1) remarks：輸出「既有銷貨單」備註 CSV（ERP銷貨單號,REMARKS），預設內容 "MM/DD ok"
   - 可選：同時輸出「進行狀態」CSV（ERP銷貨單號,進行狀態），值預設 "已購買"
2) po：依 Template-7.xlsx 產出「採購單 Excel」，每張最多 30 行，序號自動分組
   - ERP 料號 = 商城 SKU（直接填「品項編碼」）
   - 供應商 freak-j / zozo-h 對應 ERP（你給的是名稱，若 ERP 要代碼請調整）

使用開源：
- Typer（CLI）
- Rich（輸出與進度列）
- Loguru（日誌）
- openpyxl（Excel）
- Pydantic（資料驗證）
- Tenacity（保留重試機制，後續若串 API 有用）
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from typing import Optional, List, Dict

import typer
from rich.console import Console
from rich.progress import track
from loguru import logger
from tenacity import retry, stop_after_attempt, wait_exponential
from pydantic import BaseModel, Field, validator
from openpyxl import load_workbook

app = typer.Typer(help="Purchase Control Tower PRO CLI")
console = Console()

# ==== 業務參數（照你的規格） ====
EMP_CD = "TUV"       # 承辦人
WH_CD  = "100"       # 倉庫
CURRENCY = "JPY"     # 貨幣
EXCHANGE_RATE = ""   # 匯率（空白：你手動填或 ERP 自動帶）

# 供應商對照（key 為 supplier_key；value 為「ERP 供應商編碼」）
# 你提供的是名稱，如果 ERP 需要代碼（如 SUP001），請把 value 換成代碼；
# 供應商名稱顯示則在 NAME_MAP。
SUPPLIER_CODE_MAP: Dict[str, str] = {
    "freak-j": "Freaks store-傑",   # ← 若不是編碼，請改成真正 ERP 代碼
    "zozo-h":  "ZOZOTOWN-胡"
}
SUPPLIER_NAME_MAP: Dict[str, str] = {
    "freak-j": "Freaks store-傑",
    "zozo-h":  "ZOZOTOWN-胡"
}

TEMPLATE_PATH = str(Path(__file__).resolve().parents[1] / "Template-7.xlsx")
SHEET_NAME = "採購單"
TEMPLATE_COLS = [
    "日期","序號","客戶/供應商編碼","客戶/供應商名稱","承辦人","收貨倉庫","交易類型",
    "貨幣","匯率","交付日期","日本訂單號","日本貨運單號","品項編碼","品項名稱","規格",
    "日文(摘要3)","數量","單價","外幣金額","稅前價格","營業稅","訂單號","摘要"
]

MAX_LINES_PER_PO_DEFAULT = 30  # 每張採購單最多行
SERIAL_MODE_VALUES = ("group", "itemcode")  # group：1,2,3…；itemcode：序號=品項編碼（不建議）

def today_mmdd_jst() -> str:
    return datetime.now(ZoneInfo("Asia/Tokyo")).strftime("%m/%d")

def today_ymd_jst() -> str:
    return datetime.now(ZoneInfo("Asia/Tokyo")).strftime("%Y%m%d")

def esc_csv(s: str) -> str:
    s = "" if s is None else str(s)
    return '"' + s.replace('"','""') + '"' if any(c in s for c in [',','\n','"']) else s

# ==== Pydantic 資料模型 ====
class Item(BaseModel):
    supplier_key: str
    date: Optional[str] = None            # YYYYMMDD；缺省用今天
    sku: str                               # = ERP 料號
    name: Optional[str] = ""
    spec: Optional[str] = ""               # 例：奶茶/FREE（會放「摘要」「規格」）
    qty: float = Field(gt=0)
    unit_price: Optional[str] = ""
    es_order_no: Optional[str] = ""
    jp_order_no: Optional[str] = ""
    jp_waybill: Optional[str] = ""

    @validator("supplier_key")
    def _supplier_known(cls, v):
        if v not in SUPPLIER_CODE_MAP:
            raise ValueError(f"supplier_key 未定義：{v}（請在 SUPPLIER_CODE_MAP/NAME_MAP 增加）")
        return v

    @validator("date", always=True)
    def _date_default(cls, v):
        if not v: return today_ymd_jst()
        return v

# ==== 共用 ====
def read_lines(path: Path) -> List[str]:
    return [ln.strip() for ln in path.read_text(encoding="utf-8-sig").splitlines() if ln.strip()]

def write_text(path: Path, text: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8-sig")

def load_items_csv(path: Path) -> List[Item]:
    import csv
    rows: List[Item] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            try:
                rows.append(Item(**r))
            except Exception as e:
                logger.error(f"CSV 列無法解析：{r} | {e}")
    return rows

# ==== 1) 產銷貨單備註 CSV（可選輸出狀態 CSV） ====
@app.command()
def remarks(
    slip_file: Path = typer.Option(..., help="純文字檔，每行一個 ERP 銷貨單號"),
    remarks: Optional[str] = typer.Option(None, help="REMARKS 內容（預設 'MM/DD ok'）"),
    out_dir: Path = typer.Option(Path("output"), help="輸出目錄"),
    export_status_csv: bool = typer.Option(False, help="是否同時輸出『進行狀態』CSV"),
    status_value: str = typer.Option("已購買", help="進行狀態值（輸出狀態 CSV 時使用）")
):
    """
    產出：
      1) so_remarks_update_*.csv（ERP銷貨單號,REMARKS）
      2) （可選）so_status_update_*.csv（ERP銷貨單號,進行狀態）
    """
    slip_nos = read_lines(slip_file)
    if not slip_nos:
        console.print("[yellow]沒有可輸出的單號[/yellow]"); raise typer.Exit()

    if remarks is None:
        remarks = f"{today_mmdd_jst()} ok"

    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 1) 備註 CSV
    memo_path = out_dir / f"so_remarks_update_{ts}.csv"
    memo_lines = ["ERP銷貨單號,REMARKS"] + [f"{esc_csv(no)},{esc_csv(remarks)}" for no in slip_nos]
    write_text(memo_path, "\n".join(memo_lines))
    console.print(f"📄 [green]已輸出[/green] {memo_path}（{len(slip_nos)} 張單）")

    # 2) 進行狀態 CSV（可選）
    if export_status_csv:
        st_path = out_dir / f"so_status_update_{ts}.csv"
        st_lines = ["ERP銷貨單號,進行狀態"] + [f"{esc_csv(no)},{esc_csv(status_value)}" for no in slip_nos]
        write_text(st_path, "\n".join(st_lines))
        console.print(f"📄 [green]已輸出[/green] {st_path}")

    console.print("➡ 到 ECOUNT Excel/CSV 上傳，套你的映射模板即可。")

# ==== 2) 產採購單 Excel（依 Template-7.xlsx；每張最多 30 行；序號自動分組） ====
@app.command()
def po(
    items_csv: Path = typer.Option(..., help="輸入明細 CSV（欄：supplier_key,date,sku,name,spec,qty,unit_price,es_order_no,jp_order_no,jp_waybill）"),
    out_dir: Path = typer.Option(Path("output"), help="輸出目錄"),
    max_lines: int = typer.Option(MAX_LINES_PER_PO_DEFAULT, help="每張採購單最多行（預設 30）"),
    lead_days: Optional[int] = typer.Option(None, help="交付日期 = 日期 + N 天（不給則留空）"),
    serial_mode: str = typer.Option("group", help="序號模式：group | itemcode", case_sensitive=False)
):
    """
    依 Template-7.xlsx 產出可上傳的採購單 Excel（每張最多 N 行，自動分單）
    - ERP 料號 = 商城 SKU（直接填「品項編碼」）
    - 交易類型留空；貨幣 JPY；匯率空白；承辦 TUV；倉庫 100
    - 摘要/規格 填 spec（顏色/尺寸）
    """
    serial_mode = serial_mode.lower()
    if serial_mode not in SERIAL_MODE_VALUES:
        console.print("[red]serial_mode 僅能為 group 或 itemcode[/red]"); raise typer.Exit()

    items = load_items_csv(items_csv)
    if not items:
        console.print("[yellow]items_csv 沒有有效資料[/yellow]"); raise typer.Exit()

    # 依 supplier 分組
    from collections import defaultdict
    grouped: Dict[str, List[Item]] = defaultdict(list)
    for it in items:
        grouped[it.supplier_key].append(it)

    out_dir.mkdir(parents=True, exist_ok=True)

    for sup_key, arr in grouped.items():
        # 排序（日期, 訂單號, SKU）
        arr_sorted = sorted(arr, key=lambda x: (x.date, x.es_order_no or "", x.sku))
        # 分塊（每塊一張採購單）
        chunks = [arr_sorted[i:i+max_lines] for i in range(0, len(arr_sorted), max_lines)]

        # 載入模板
        wb = load_workbook(TEMPLATE_PATH); ws = wb[SHEET_NAME]
        header = [c.value for c in ws[1]]
        if header[:len(TEMPLATE_COLS)] != TEMPLATE_COLS:
            console.print("[red]Template-7.xlsx 欄位與預期不符[/red]"); raise typer.Exit()

        write_row = 2
        serial = 1   # group 模式：同一張單，同一序號（<=4）
        for chunk in track(chunks, description=f"[cyan]寫入 {sup_key}[/cyan]"):
            for it in chunk:
                ymd = it.date or today_ymd_jst()
                cust_cd  = SUPPLIER_CODE_MAP[sup_key]   # 若需真正 ERP 代碼，請把 map 改成代碼
                cust_des = SUPPLIER_NAME_MAP[sup_key]
                emp_cd, wh_cd = EMP_CD, WH_CD
                io_type = ""                 # 交易類型留空
                curr, exrate = CURRENCY, EXCHANGE_RATE
                time_date = ""
                if lead_days is not None:
                    dt = datetime.strptime(ymd, "%Y%m%d"); time_date = (dt + timedelta(days=int(lead_days))).strftime("%Y%m%d")
                prod_cd, prod_des = it.sku, it.name or ""
                spec, qty, unit_price = it.spec or "", it.qty, it.unit_price or ""
                amt_f = amt = vat = ""
                es_no, jp_no, jp_way = it.es_order_no or "", it.jp_order_no or "", it.jp_waybill or ""
                if serial_mode=="itemcode":
                    ser_no = str(prod_cd)[:4]  # 風險：一行一單、可能超 4 被截斷
                else:
                    ser_no = str(serial)       # 建議：group 模式

                row = [
                    ymd, ser_no, cust_cd, cust_des, emp_cd, wh_cd, io_type,
                    curr, exrate, time_date, jp_no, jp_way, prod_cd, prod_des,
                    spec, "", qty, unit_price, amt_f, amt, vat, es_no, spec
                ]
                for ci, val in enumerate(row, start=1):
                    ws.cell(row=write_row, column=ci, value=val)
                write_row += 1

            serial = serial + 1 if serial < 9999 else 1

        out_path = out_dir / f"po_upload_{sup_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(out_path)
        console.print(f"📦 產出 {out_path}（{len(arr_sorted)} 行，分 {len(chunks)} 張單；每張 ≤ {max_lines} 行）")

    console.print("➡ 到 ECOUNT Excel 上傳（採購單），套你的 Template-7 對映即可。")

if __name__ == "__main__":
    app()

